"""Update Pruebas xlsx with test results (new columns P, Q, R)."""
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Pruebas desarrollo monitor v1.xlsx"
RESULTS = ROOT / "test_run_results.json"

ESTADO_LABEL = {
    "01": "01 Pendiente", "02": "02 Validado", "03": "03 Rechazado",
    "04": "04 Por Rechazar", "05": "05 Error", "06": "06 Contabilizado",
}

results = {int(r["case"]): r for r in json.loads(RESULTS.read_text(encoding="utf-8"))}

wb = load_workbook(XLSX)
ws = wb["Servicios"]

# New headers (cols P, Q, R, S)
headers = {
    "P1": "Estado obtenido",
    "Q1": "Log Procesamiento",
    "R1": "Resultado",
    "S1": "Notas ejecucion",
}
header_font = Font(bold=True)
header_fill = PatternFill("solid", start_color="D9E1F2")
for coord, val in headers.items():
    c = ws[coord]
    c.value = val
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

pass_fill = PatternFill("solid", start_color="C6EFCE")
fail_fill = PatternFill("solid", start_color="FFC7CE")
skip_fill = PatternFill("solid", start_color="FFEB9C")
pass_font = Font(color="006100", bold=True)
fail_font = Font(color="9C0006", bold=True)
skip_font = Font(color="9C6500", bold=True)

# Column widths
ws.column_dimensions["P"].width = 16
ws.column_dimensions["Q"].width = 55
ws.column_dimensions["R"].width = 10
ws.column_dimensions["S"].width = 45

for row_idx in range(2, ws.max_row + 1):
    num_cell = ws.cell(row=row_idx, column=1).value
    if isinstance(num_cell, str) and num_cell.startswith("="):
        # Formula — case number = row_idx - 1
        num = row_idx - 1
    else:
        num = int(num_cell) if num_cell is not None else None
    if num is None:
        continue
    r = results.get(num)
    if not r:
        continue

    if r.get("skipped"):
        ws.cell(row=row_idx, column=16, value="-").fill = skip_fill
        ws.cell(row=row_idx, column=17, value=f"(Skip: {r.get('reason','')})").fill = skip_fill
        cell = ws.cell(row=row_idx, column=18, value="SKIP")
        cell.fill = skip_fill
        cell.font = skip_font
        cell.alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=19, value="Caso marcado 'Pendiente' en plan, sin resultado esperado.").fill = skip_fill
        continue

    estado_label = ESTADO_LABEL.get(r.get("estado"), r.get("estado") or "?")
    ws.cell(row=row_idx, column=16, value=estado_label)
    ws.cell(row=row_idx, column=17, value=r.get("log_proc") or "").alignment = Alignment(wrap_text=True, vertical="top")

    pass_ = r.get("pass")
    if pass_ is True:
        cell = ws.cell(row=row_idx, column=18, value="PASS")
        cell.fill = pass_fill
        cell.font = pass_font
    else:
        cell = ws.cell(row=row_idx, column=18, value="FAIL")
        cell.fill = fail_fill
        cell.font = fail_font
    cell.alignment = Alignment(horizontal="center")

    notas = []
    if r.get("oc"):
        notas.append(f"OC SAP: {r['oc']}")
    if r.get("hes"):
        notas.append(f"HES SAP: {r['hes']}")
    if r.get("doc_fact"):
        notas.append(f"Doc fact: {r['doc_fact']}")
    if not pass_:
        # Diagnose
        log = (r.get("log_proc") or "").lower()
        exp = (r.get("expected") or "").lower()
        if "no corresponde al proveedor" in log and "hes" in exp:
            notas.append("La validacion de OC frena antes de evaluar la HES (dato maestro de OC en SAP no usa proveedor 76948695-K).")
        elif "se debe generar factura" in exp:
            notas.append("El motor freno antes de validar monto — revisar dato maestro de la OC/HES en SAP.")
    if notas:
        ws.cell(row=row_idx, column=19, value=" | ".join(notas)).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row_idx].height = 60

# Add a summary footer
last_row = ws.max_row + 2
ws.cell(row=last_row, column=1, value=f"Ejecutado: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True)
total = sum(1 for r in results.values() if not r.get("skipped"))
passed = sum(1 for r in results.values() if r.get("pass") is True)
failed = sum(1 for r in results.values() if r.get("pass") is False and not r.get("skipped"))
skipped = sum(1 for r in results.values() if r.get("skipped"))
ws.cell(row=last_row + 1, column=1, value=f"Resumen: {passed} PASS / {failed} FAIL / {skipped} SKIP (de {total + skipped})").font = Font(bold=True)

wb.save(XLSX)
print(f"Actualizado: {XLSX}")
print(f"Resumen: {passed} PASS / {failed} FAIL / {skipped} SKIP (de {total + skipped})")
