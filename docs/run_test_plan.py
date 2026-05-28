"""
Orquestador para correr el plan de pruebas del Monitor DTE.
- Lee el xlsx de plan de pruebas.
- Para cada caso ejecutable: POST IngestFromXml -> POST IngestFromSii -> GET registro.
- Compara LogProcesamiento con el Resultado esperado.
- Escribe resultado obtenido y estado (PASS/FAIL) en la misma planilla.
"""
import json
import os
import sys
import re
import argparse
from datetime import date
from pathlib import Path

import requests
from openpyxl import load_workbook

URL = "https://my419159-api.s4hana.cloud.sap/sap/opu/odata4/sap/zui_dte_monitor_o4/srvd/sap/zui_dte_monitor/0001"
USER = "COM_MONITOR"
PWD = "5Nt[#L4>K2/8xAy7#6+nMFNlsGC-x9Gb~5&b]%o]"
NS = "com.sap.gateway.srvd.zui_dte_monitor.v0001"

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "Pruebas desarrollo monitor v1.xlsx"
XML_DIR = ROOT / "XML"
RESULTS_JSON = ROOT / "test_run_results.json"


def make_session():
    s = requests.Session()
    s.auth = (USER, PWD)
    s.headers.update({"Accept": "application/json"})
    r = s.head(URL + "/", headers={"x-csrf-token": "fetch"}, timeout=30)
    r.raise_for_status()
    csrf = r.headers.get("x-csrf-token")
    if not csrf:
        raise RuntimeError(f"No CSRF token. Headers: {dict(r.headers)}")
    s.headers["x-csrf-token"] = csrf
    s.headers["Content-Type"] = "application/json"
    return s


def post_action(s, name, body):
    url = f"{URL}/DteMonitor/{NS}.{name}"
    r = s.post(url, data=json.dumps(body), timeout=60)
    return r


def get_record(s, tipo, folio, proveedor):
    key = f"DteMonitor(TipoDte='{tipo}',Folio='{folio}',Proveedor='{proveedor}')"
    fields = "TipoDte,Folio,Proveedor,Sociedad,Estado,LogProcesamiento,FechaRecepcionSii,OrdenCompra,HojaEntradaServicio,EntradaMercancia,DocumentoFacturaSap,AnioFacturaSap"
    r = s.get(f"{URL}/{key}?$select={fields}", timeout=30)
    return r


def delete_record(s, tipo, folio, proveedor):
    key = f"DteMonitor(TipoDte='{tipo}',Folio='{folio}',Proveedor='{proveedor}')"
    return s.delete(f"{URL}/{key}", timeout=30)


def reprocesar(s, tipo, folio, proveedor):
    url = (
        f"{URL}/DteMonitor(TipoDte='{tipo}',Folio='{folio}',Proveedor='{proveedor}')"
        f"/{NS}.Reprocesar"
    )
    r = s.post(url, data="{}", timeout=60)
    return r


def read_xml(name):
    p = XML_DIR / f"{name}.xml"
    return p.read_text(encoding="ISO-8859-1")


ESTADO_LABEL = {
    "01": "Pendiente",
    "02": "Validado",
    "03": "Rechazado",
    "04": "Por Rechazar",
    "05": "Error",
    "06": "Contabilizado",
}


def run_case(s, case):
    """case = dict with: num, descripcion, sociedad_rut, proveedor_rut, folio,
       xml_name, monto_total, fecha_doc, moneda, expected"""
    log = []
    folio = str(case["folio"])
    tipo = "033"
    proveedor = case["proveedor_rut"]
    sociedad = case["sociedad_rut"] or "00000000-0"

    # 1) Read XML
    try:
        xml_data = read_xml(case["xml_name"])
    except FileNotFoundError as e:
        return {
            "case": case["num"],
            "error": f"XML no encontrado: {e}",
            "estado": None,
            "log_proc": None,
            "expected": case["expected"],
            "pass": False,
        }

    # 1.5) Borrar registro pre-existente para que el ingest parta limpio
    #      (IngestFromXml es solo-CREATE y XmlData es readonly:update, así que
    #       sin borrar antes el XML viejo persiste — ver caso 17).
    rd = delete_record(s, tipo, folio, proveedor)
    log.append(f"DELETE HTTP {rd.status_code}")

    # 2) IngestFromXml
    r1 = post_action(s, "IngestFromXml", {"XmlData": xml_data})
    log.append(f"IngestFromXml HTTP {r1.status_code}")
    try:
        body1 = r1.json()
    except Exception:
        body1 = r1.text[:500]
    log.append(f"  body: {str(body1)[:300]}")

    # 3) IngestFromSii (use receptor RUT as Sociedad as XML extracted it)
    sii_body = {
        "TipoDte": tipo,
        "Folio": folio,
        "Proveedor": proveedor,
        "Sociedad": sociedad,
        "FechaRecepcionSii": date.today().isoformat(),
        "FechaDocumento": case.get("fecha_doc") or "2026-05-25",
        "Moneda": case.get("moneda") or "CLP",
        "MontoTotal": float(case.get("monto_total") or 0),
    }
    r2 = post_action(s, "IngestFromSii", sii_body)
    log.append(f"IngestFromSii HTTP {r2.status_code}")
    try:
        body2 = r2.json()
    except Exception:
        body2 = r2.text[:500]
    log.append(f"  body: {str(body2)[:300]}")

    # 4) GET record
    r3 = get_record(s, tipo, folio, proveedor)
    log.append(f"GET HTTP {r3.status_code}")
    rec = None
    if r3.status_code == 200:
        rec = r3.json()
        # If state is final and was already validated previously, force Reprocesar to get fresh log
        if rec.get("Estado") in ("02", "03", "04", "06"):
            # Force re-validation in case record existed from previous run
            r4 = reprocesar(s, tipo, folio, proveedor)
            log.append(f"Reprocesar HTTP {r4.status_code}")
            r5 = get_record(s, tipo, folio, proveedor)
            log.append(f"GET-2 HTTP {r5.status_code}")
            if r5.status_code == 200:
                rec = r5.json()

    estado = rec.get("Estado") if rec else None
    log_proc = (rec.get("LogProcesamiento") or "") if rec else ""

    # 5) Compare
    expected = (case["expected"] or "").strip()
    pass_ = compare(expected, log_proc, estado, case)

    return {
        "case": case["num"],
        "folio": folio,
        "estado": estado,
        "estado_label": ESTADO_LABEL.get(estado, "?"),
        "log_proc": log_proc,
        "oc": rec.get("OrdenCompra") if rec else None,
        "hes": rec.get("HojaEntradaServicio") if rec else None,
        "doc_fact": rec.get("DocumentoFacturaSap") if rec else None,
        "expected": expected,
        "pass": pass_,
        "trace": log,
    }


def compare(expected, log_proc, estado, case):
    """Heuristic comparator. expected contains the keyword/message; we check log_proc contains the essence."""
    if not expected:
        return None
    exp_norm = re.sub(r"[‘’“”\"']", "", expected).strip().lower()
    log_norm = re.sub(r"[‘’“”\"']", "", (log_proc or "")).strip().lower()

    if "se debe generar factura" in exp_norm:
        # Expect validation OK / Validado / facturado
        return estado in ("02", "06")
    # Extract the message between quotes if present
    m = re.search(r"mensaje\s+[‘’\"']?(.+?)[‘’\"']?$", exp_norm)
    needle = m.group(1).strip(" .'\"") if m else exp_norm
    return needle in log_norm


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", type=str, default="", help="Comma-separated case numbers to run (e.g. 1,2,3)")
    ap.add_argument("--smoke", action="store_true", help="Only verify connectivity")
    args = ap.parse_args()

    print("Conectando al tenant…", flush=True)
    s = make_session()
    print(f"  CSRF token: {s.headers['x-csrf-token'][:8]}…", flush=True)

    if args.smoke:
        print("Smoke OK.")
        return

    # Load cases from xlsx
    wb = load_workbook(XLSX, data_only=True)
    ws = wb["Servicios"]
    cases = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        num, desc, rut_soc, soc, prov, rut_prov, folio, oc, hes, neto, iva, total, xml_name, expected, obs = (
            row + (None,) * (15 - len(row))
        )[:15]
        if not folio or not xml_name:
            continue
        # Col A may be a formula with empty cache — derive from row index
        if num is None:
            num = row_idx - 1
        cases.append({
            "num": num,
            "descripcion": desc,
            "sociedad_rut": rut_soc,
            "sociedad": soc,
            "proveedor": prov,
            "proveedor_rut": rut_prov,
            "folio": folio,
            "oc": oc,
            "hes": hes,
            "monto_neto": neto,
            "iva": iva,
            "monto_total": total,
            "xml_name": str(xml_name).strip(),
            "expected": expected,
            "observacion": obs,
        })

    if args.only:
        wanted = set(int(x) for x in args.only.split(",") if x.strip())
        cases = [c for c in cases if c["num"] is not None and int(c["num"]) in wanted]

    results = []
    for c in cases:
        # Skip case 9 (marked 'Pendiente' — no expected result, no OC/HES)
        if not c["expected"]:
            print(f"[{c['num']:>2}] {c['descripcion']!r} — SKIP (sin expected)")
            results.append({"case": c["num"], "skipped": True, "reason": "sin expected", "pass": None})
            continue
        print(f"[{c['num']:>2}] {c['descripcion']!r}", flush=True)
        try:
            r = run_case(s, c)
        except Exception as e:
            r = {"case": c["num"], "error": repr(e), "pass": False}
        results.append(r)
        status = "PASS" if r.get("pass") else ("SKIP" if r.get("skipped") else "FAIL")
        log_excerpt = (r.get("log_proc") or r.get("error") or "")[:120]
        print(f"     -> {status} estado={r.get('estado')} log={log_excerpt!r}", flush=True)

    RESULTS_JSON.write_text(json.dumps(results, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"\nResultados escritos a {RESULTS_JSON}")


if __name__ == "__main__":
    main()
